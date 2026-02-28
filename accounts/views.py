from django.shortcuts import render
from rest_framework.response import Response
from rest_framework import generics, permissions, status
from rest_framework.permissions import AllowAny
from .models import User, UserVerification
from .serializers import AdminCreateUserSerializer, UserRegisterSerializer, VerificationSerializer, VerificationUpdateSerializer
from django.utils import timezone
from django.db import transaction
from rest_framework.permissions import IsAuthenticated
from .serializers import UserProfileSerializer

from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated
from .serializers import UserUpdateSerializer
from openpyxl import load_workbook
from rest_framework import status
from .models import SemesterResult, Student, Faculty
from .utils import send_result_notification
from .serializers import SemesterResultSerializer, FacultyListSerializer, FacultyDetailSerializer
from rest_framework.views import APIView
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework import serializers
from .permissions import IsCustomAdmin
from .serializers import (
    UserAdminListSerializer,
    UserAdminDetailSerializer,
    UserAdminUpdateSerializer,
)
from .pagination import AdminUserPagination


# Create your views here.
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegisterSerializer
    permission_classes = [AllowAny]
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class UserListView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegisterSerializer
    permission_classes = [AllowAny]


class VerificationStatusView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        verifications = request.user.verifications.all()
        data = [
            {
                "type": v.verification_type,
                "verified": v.is_verified,
                "submitted_at": v.submitted_at
            }
            for v in verifications
        ]
        return Response(data)
    

class PendingVerificationListView(generics.ListAPIView):
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        pending = UserVerification.objects.filter(is_verified=False)
        data = [
            {
                "id":v.id,
                "user":v.user.email,
                "role":v.user.role,
                "type":v.verification_type,
                "submitted_at":v.submitted_at,
                "document":v.document.url
            }
            for v in pending
        ]
        return Response(data)

class ReviewVerificationView(APIView):
    permission_classes = [permissions.IsAdminUser]
    
    def patch(self, request, pk):
        try:
            verification = UserVerification.objects.get(pk=pk)
        except UserVerification.DoesNotExist:
            return Response(
                {"detail": "Verification not found"}, status=status.HTTP_404_NOT_FOUND
            )
            
        if verification.is_verified:
            return Response(
                {"detail": "Verification already reviewed"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        approve = request.data.get("approve",False)
        remarks = request.data.get("remarks","")
        
        with transaction.atomic():
            verification.is_verified = approve
            verification.reviewed_at = timezone.now()
            verification.reviewed_by = request.user
            verification.remarks = remarks
            verification.save()

            # Approve user only if all verifications are approved
            user = verification.user
            if approve and not user.verifications.filter(is_verified=False).exists():
                user.is_approved = True
                user.save()
            
        return Response({"status":"updated"})
    



class ProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        serializer = UserProfileSerializer(request.user, context={'request': request})
        return Response(serializer.data)
    
    
class ProfileUpdateView(generics.UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserUpdateSerializer

    def get_object(self):
        return self.request.user

    def patch(self, request, *args, **kwargs):
        user = self.get_object()
        serializer = self.get_serializer(user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

class FacultyListView(generics.ListAPIView):
    queryset = Faculty.objects.all().select_related('profile__user')
    serializer_class = FacultyListSerializer
    permission_classes = [AllowAny]
    
    # def get_queryset(self):
    #     return super().get_queryset().filter(profile__user__role='faculty')
    def get_queryset(self):
        from django.db.models import Case, When, Value, IntegerField

        designation_order = {
            'Professor': 1,
            'Associate Professor': 2,
            'Assistant Professor': 3,
            'Lecturer': 4,
        }

        return Faculty.objects.select_related('profile__user').annotate(
            designation_priority=Case(
                *[When(designation=desig, then=Value(priority)) for desig, priority in designation_order.items()],
                default=Value(999),
                output_field=IntegerField()
            )
        ).order_by('designation_priority', 'profile__user__last_name', 'profile__user__first_name')
    

class FacultyDetailView(generics.RetrieveAPIView):
    queryset = Faculty.objects.all().select_related('profile__user')
    serializer_class = FacultyDetailSerializer
    permission_classes = [AllowAny]
    lookup_field = 'id'
    
    

# ======================================
# ADMIN USER LIST (with pagination, filters, search, ordering)
# ======================================
class AdminUserListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    serializer_class = UserAdminListSerializer
    queryset = User.objects.all().select_related('profile')
    
    # Filters & Search & Ordering
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'is_approved']               # ?role=faculty&is_approved=true
    search_fields = ['email', 'username', 'first_name', 'last_name']  # ?search=akram
    ordering_fields = ['date_joined', 'role', 'is_approved', 'last_name']
    ordering = ['-date_joined']                              # default newest first

    # Pagination
    pagination_class = AdminUserPagination
    

# ======================================
# ADMIN USER DETAIL / EDIT / DELETE
# ======================================
class AdminUserDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    queryset = User.objects.all().select_related('profile')
    lookup_field = 'id'

    def get_serializer_class(self):
        # Use update serializer for PUT/PATCH, detail for GET
        if self.request.method in ['PUT', 'PATCH']:
            return UserAdminUpdateSerializer
        return UserAdminDetailSerializer

    def perform_update(self, serializer):
        # Optional: extra logic before save (e.g. prevent changing own role to non-admin)
        if serializer.instance == self.request.user and serializer.validated_data.get('role') != 'admin':
            raise DjangoValidationError("Admins cannot demote themselves.")
        serializer.save()

    def perform_destroy(self, instance):
        # Safety: prevent deleting the last admin
        if instance.role == 'admin' and User.objects.filter(role='admin').count() == 1:
            raise DjangoValidationError("Cannot delete the last admin user.")
        instance.delete()
        


# ======================================
# QUICK APPROVE ENDPOINT (optional but very useful)
# ======================================
class AdminApproveUserView(generics.UpdateAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    serializer_class = serializers.Serializer
    queryset = User.objects.all()
    lookup_field = 'id'

    def update(self, request, *args, **kwargs):
        user = self.get_object()

        # Get data from request body (you can send from frontend)
        is_approved = request.data.get('is_approved', None)
        remarks = request.data.get('remarks', None)

        if is_approved is None:
            return Response({"detail": "is_approved field is required"}, status=400)

        if user.is_approved == is_approved:
            return Response({"detail": f"User is already {'approved' if is_approved else 'not approved'}"}, status=400)

        user.is_approved = is_approved
        user.is_active = is_approved
        user.save()

        # Send email
        send_account_approval_email(
            user=user,
            is_approved=is_approved,
            remarks=remarks
        )

        return Response({
            "detail": f"User {'approved' if is_approved else 'approval revoked'} successfully"
        }, status=200)

class AdminCreateUserView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    serializer_class = AdminCreateUserSerializer


# List verifications for a specific user
class AdminUserVerificationListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    serializer_class = VerificationSerializer

    def get_queryset(self):
        user_id = self.kwargs['user_id']
        return UserVerification.objects.filter(user_id=user_id)
    

class AdminUserVerificationListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    serializer_class = VerificationSerializer

    def get_queryset(self):
        user_id = self.kwargs['user_id']
        return UserVerification.objects.filter(user_id=user_id).order_by('-submitted_at')


class AdminUserVerificationDetailView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated, IsCustomAdmin]
    serializer_class = VerificationUpdateSerializer
    lookup_field = 'id'

    def get_queryset(self):
        user_id = self.kwargs['user_id']
        return UserVerification.objects.filter(user_id=user_id)

    def perform_update(self, serializer):
        serializer.save(
            reviewed_by=self.request.user,
            reviewed_at=timezone.now()
        )
        
class AdminResultUploadView(APIView):
    permission_classes = [IsCustomAdmin]
    def post(self,request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error":"No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = load_workbook(file)
            ws = wb.active
            success = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    student_id = str(row[0]).strip()
                    semester = str(row[1]).strip()
                    sgpa = float(row[2])
                    cgpa = float(row[3])
                    status = str(row[4]).strip().lower()
                    
                    student = Student.objects.get(student_id=student_id)
                    
                    SemesterResult.objects.update_or_create(
                        student=student,
                        semester=semester,
                        defaults={
                            'sgpa':sgpa,
                            'cgpa':cgpa,
                            'result_status':status,
                            'uploaded_by':request.user
                        }
                    )
                    
                    send_result_notification(student.user,semester,sgpa,cgpa,status)
                    success +=1
                except Student.DoesNotExist:
                    errors.append(f"Student with ID {student_id} not found!")
                except Exception as e:
                    errors.append(f"Error in row {str(e)}")
            return Response({
                "message": f"Successfully processed {success} results.",
                "errors": errors[:10]
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error":str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        
class StudentResultsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        print(f"User: {request.user.email} is accessing student-results")

        try:
            profile = request.user.profile
            print(f"Profile found: {profile}")

            student = profile.student
            print(f"Student found: {student.student_id}")

            results = SemesterResult.objects.filter(student=student)
            print(f"Found {results.count()} results")

            serializer = SemesterResultSerializer(results, many=True)
            return Response(serializer.data)

        except AttributeError as e:
            print(f"AttributeError: {str(e)}")
            return Response({"detail": "Missing profile or student relation"}, status=404)

        except Student.DoesNotExist:
            print("Student.DoesNotExist")
            return Response({"detail": "No student profile linked to this user"}, status=404)

        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            return Response({"detail": f"Server error: {str(e)}"}, status=500)