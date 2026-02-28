from rest_framework import serializers
from .models import Notice
from django.urls import reverse
from accounts.models import Student, SemesterResult
from accounts.utils import send_result_notification
from openpyxl import load_workbook
import pytesseract
from PIL import Image
import io

class NoticeSerializer(serializers.ModelSerializer):
    download_url = serializers.SerializerMethodField()
    category_display = serializers.CharField(source='get_category_display', read_only=True)
    class Meta:
        model = Notice
        fields = ['id', 'title', 'category', 'category_display', 'date', 'download_url']

    def get_download_url(self, obj):
        request = self.context.get('request')
        if request and request.user.is_authenticated:
            return request.build_absolute_uri(reverse('notice-download', kwargs={'pk': obj.id}))
        return None


class NoticeCreateSerializer(serializers.ModelSerializer):
    result_excel = serializers.FileField(required=False, write_only=True)

    class Meta:
        model = Notice
        fields = ['title', 'category', 'pdf_file', 'result_excel']

    def create(self, validated_data):
        result_excel = validated_data.pop('result_excel', None)
        notice = super().create(validated_data)

        if result_excel and notice.category == 'exam':
            self.process_excel(result_excel, notice)

        return notice

    def process_excel(self, excel_file, notice):
        from openpyxl import load_workbook

        wb = load_workbook(excel_file)
        ws = wb.active
        success_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            try:
                student_id = str(row[1]).strip()
                sgpa = float(row[5])
                cgpa = float(row[6])
                status = str(row[7]).strip().lower()

                semester = self.context['request'].data.get('semester')
                if not semester:
                    raise ValueError("Semester is required for exam notices")

                student = Student.objects.get(student_id=student_id)

                SemesterResult.objects.update_or_create(
                    student=student,
                    semester=semester,
                    defaults={
                        'sgpa': sgpa,
                        'cgpa': cgpa,
                        'result_status': status,
                        'uploaded_by': self.context['request'].user
                    }
                )

                send_result_notification(
                    student.profile.user, 
                    semester, 
                    sgpa, 
                    cgpa, 
                    status
                )
                success_count += 1

            except Student.DoesNotExist:
                errors.append(f"Row {row_idx}: Student ID {student_id} not found")
            except ValueError as ve:
                errors.append(f"Row {row_idx}: {str(ve)}")
            except AttributeError as ae:
                errors.append(f"Row {row_idx}: Relation error - {str(ae)}")
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        print(f"Processed {success_count} students from Excel. Errors: {errors}")
        if errors:
            print("Full errors:", errors)